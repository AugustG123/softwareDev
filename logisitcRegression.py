import torch
import numpy as np
import openpyxl as xl
import csv


#def imported(filename):
    wb = xl.open_workbook(filename)
    sheet = wb.active
    all_data=np.empty(sheet.max_row, sheet.max_col)

    for row in range(1,sheet.max_row+1):
        for data in range(0, sheet.max_col+1):
            all_data[row][col] = sheet.cell(row,col)
    return all_data
            


def arrayMultiplication(constant, array):
    for i in range(len(array)):
        array[i]*=constant
    return array


#array1 and array2 need to have the same size
def arrayAddition(array1, array2):
    array_final=[]
    for i in range(len(array1)):
        array_final.append(array1[i] + array2[i])
    return array_final


def dot_product(array1, array2):
    total = 0.0
    for i in range(len(array1)):
        total += (float)(array1[i]*array2[i])
    return total


def logti_function(z):
    return (float) (1/((float)(1+e**(-z))))


def average_cost(weights,y): #damage level =1,2,3 or 4
        total_cost = 0
        for i in range(len(y)):
            total_cost+=-log(abs((float)y[i] - logti_function(dot_product(weights, alldata[i]))), 10)
        return total_cost/len(y)


def random_list(length):
    array = []
    for i in range(length):
        array.append(100.0 * random.random())
    return array

def delete_a_column(array, row): #input row in index form  
    for i in range(len(array)):
        array[i].pop(row)
    return array


class LogisticRegressionModel:

    def __init__(self, data, num_iterations, learning_rate, model_number, y):
        self.data = data
        self.data_transposed = alldata.transpose(alldata, len(alldata[0]), len(alldata))
        self.num_iterations = num_iterations
        self.learning_rate = learning_rate
        self.model_number = model_number
        self.y = y #output list
        self.weights = np.ones(len(data[0]))


    def gradient_descent():
        for i in range(100):
            new_weights = random_list(len(self.data[0]))
            for i in range(self.num_iterations):
                all_predictions = []
                for j in range(len(self.data)):
                    all_predictions.append(logti_function(dot_product(alldata[j],new_weights)))
                error_array = arrayAddition(all_predictions, arrayMultiplication(-1,self.y))
        
                gradients = []
                for k in range(len(self.alldata_transposed)):
                    gradients.append(dot_product(self.alldata_transposed[k], error_array))
                gradients = arrayMultiplication((1/len(self.data)),gradients)
        
                new_weights = arrayAddition(new_weights, arrayMultiplication(-1*self.learning_rate, gradients))
            if average_cost(new_weights,self.y) >= average_cost(self.weights, self.y):
                self.weights = new_weights
            

reader = csv.reader(open('Biomarkers-Blood Cells-EPCs for analysis ML.csv'))
inputList = []
targetList = []
all_Nan = []
o = 0
for row in reader:
  li = []
  row_Nan = []
  if o > 216:
    break
  if o == 0:
    o += 1
    continue
  index = -2
  for value in row:
    try:
      li.append(float(value))
      row_Nan.append(0)
    except:
      li.append(None)
      row_Nan.append(1)
    index += 1
  try:
    inputList.append(li[2:])
    targetList.append(li[1])
    all_Nan.append(row_Nan[2:])
  except:
    print(li)
  o += 1

nan_per_parameter = []
for l in range(len(all_Nan[0])):
    total_Nan = 0
    for p in range(len(all_Nan)):
        total_Nan += all_Nan[p][l]
    nan_per_parameter.append(total_Nan)

num_columns_deleted = 0
for r in range(len(nan_per_parameter)):
    if nan_per_parameter[r] >= .1 * len(all_Nan):
        inputList = delete_a_column(inputList, r-num_colums_deleted)
        num_columns_deleted += 1
        



train_data=imported("pvalue_test2.xlsx")
test_data = #some method

damage_level_train_list = #some method
damage_level_test_list = #some method

y1 = []
for i in range(len(train_data)):
    if damage_level_train_list[i] == 1:
        y1.append(1)
    else:
        y1.append(0)
model1 = LogisticRegressionModel(train_data, num_iterations = 1000, learning_rate = .05, model_number=1, y= y1)
model1.gradient_descent()

y2 = []
for i in range(len(train_data)):
    if damage_level_train_list[i] == 2:
        y2.append(1)
    else:
        y2.append(0)
model2 = LogisticRegressionModel(train_data, num_iterations = 1000, learning_rate = .05, model_number=2, y= y2)
model2.gradient_descent()

y3 = []
for i in range(len(train_data)):
    if damage_level_train_list[i] == 3:
        y3.append(1)
    else:
        y3.append(0)
model3 = LogisticRegressionModel(train_data, num_iterations = 1000, learning_rate = .05, model_number=3, y= y3)
model3.gradient_descent()

y4 = []
for i in range(len(train_data)):
    if damage_level_train_list[i] == 4:
        y2.append(1)
    else:
        y2.append(0)
model4 = LogisticRegressionModel(train_data, num_iterations = 1000, learning_rate = .05, model_number=4, y= y4)
model4.gradient_descent()

number_correct = 0
for m in range(len(test_data))
    prob1 = logti_function(dot_product(model1.weights, test_data[m]))
    prob2 = logti_function(dot_product(model2.weights, test_data[m]))
    prob3 = logti_function(dot_product(model3.weights, test_data[m]))
    prob4 = logti_function(dot_product(model4.weights, test_data[m]))
    predict_index = 0
    max_prob = prob1
    predictions = [prob1, prob2, prob3, prob4]
    for index in range(1,len(predictions)):
        if predictions[index] > max_prob:
            predict_index = index
            max_prob = predictions[index]
    prediction = predict_index + 1
    if damage_level_test_list[m] == prediction:
        number_correct += 1

print(number_correct/len(test_data))
